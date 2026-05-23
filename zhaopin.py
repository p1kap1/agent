"""智联招聘数据抓取模块

API 端点:
- 已投递: /schedule/feedback?status=send
- 我的收藏: /schedule/feedback?status=collect  
- 职位推荐: /position/recommend

认证: at + rt token (Query参数) + Cookie + 特殊 Headers
"""

import json
import os
import time
import uuid
import random
from datetime import date, datetime

import requests

from config import DATA_DIR

# ---- 配置 ----
ZHAOPIN_BASE = "https://fe-api.zhaopin.com/c/i"

TAB_CONFIG = {
    "智联-已投递": {
        "endpoint": "/schedule/feedback",
        "params": {"status": "send", "storeViewCount": "false"},
    },
    "智联-收藏": {
        "endpoint": "/schedule/feedback",
        "params": {"status": "collect", "storeViewCount": "false"},
        "api_base": "https://cgate.zhaopin.com/behavior/userforward/listPositionCollect",
    },
    "智联-推荐": {
        "endpoint": "/position/recommend",
        "params": {},
    },
}


def _load_zhaopin_config() -> dict:
    users_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "users.json")
    if not os.path.exists(users_file):
        return {}
    with open(users_file, "r", encoding="utf-8") as f:
        users = json.load(f)
    active = users.get("active_user", "default")
    user = users.get("users", {}).get(active, {})
    return user.get("zhaopin", {})


def _get_headers() -> dict:
    cfg = _load_zhaopin_config()
    cookie_data = cfg.get("cookie_json", "{}")
    if isinstance(cookie_data, str):
        cookie_data = json.loads(cookie_data) if cookie_data.strip() else {}

    cookie_str = cookie_data.get("cookies", "") if isinstance(cookie_data, dict) else ""
    ua = cookie_data.get("userAgent", "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36") if isinstance(cookie_data, dict) else ""

    # 从 cookie 提取 at 和 rt
    cookies = {}
    for pair in cookie_str.split("; "):
        if "=" in pair:
            k, v = pair.split("=", 1)
            cookies[k] = v

    return {
        "User-Agent": ua,
        "Cookie": cookie_str,
        "x-zp-client-id": cookies.get("x-zp-client-id", ""),
        "x-zp-device-sn": cookies.get("x-zp-device-sn", ""),
        "x-zp-page-request-id": f"{uuid.uuid4()}-{int(time.time()*1000)}-{random.randint(100,999)}",
    }, cookies


def _request_api(endpoint: str, extra_params: dict = None) -> dict:
    headers, cookies = _get_headers()
    at_token = cookies.get("at", "")
    rt_token = cookies.get("rt", "")

    if not at_token:
        raise PermissionError("请在 users.json 中配置 zhaopin.cookie_json")

    params = {"at": at_token, "rt": rt_token}
    if extra_params:
        params.update(extra_params)

    url = f"{ZHAOPIN_BASE}{endpoint}"
    resp = requests.get(url, params=params, headers=headers, timeout=20)

    if resp.status_code in (401, 403):
        raise PermissionError("智联登录态已过期，请更新 Cookie")
    resp.raise_for_status()
    data = resp.json()

    code = data.get("code")
    if code is not None and code != 200 and code != 0:
        msg = data.get("message", "未知错误")
        raise RuntimeError(f"智联 API 错误 (code={code}): {msg}")

    return data


def _parse_item(item: dict, status: str) -> dict:
    """解析智联职位数据为统一格式"""
    staff = item.get("staffCard") or {}

    # 计算 happen_time：已投递/收藏的 displayDate 才是投递时间，不是 jobPostingTime
    display_date = item.get("displayDate", "")
    happen_time = ""
    if display_date:
        from datetime import datetime as _dt, timedelta as _td
        if "今天" in display_date:
            happen_time = str(int(_dt.now().timestamp() * 1000))
        elif "昨天" in display_date:
            yesterday = _dt.now() - _td(days=1)
            happen_time = str(int(yesterday.timestamp() * 1000))
        else:
            # 尝试解析 "YYYY-MM-DD" 或 "MM-DD" 格式
            try:
                ds = display_date[:10]
                if len(ds) >= 10:
                    d = _dt.strptime(ds, "%Y-%m-%d")
                else:
                    d = _dt.strptime(f"{_dt.now().year}-{ds}", "%Y-%m-%d")
                happen_time = str(int(d.timestamp() * 1000))
            except Exception:
                pass
    # 兜底：推荐用 jobPostingTime，已投递/收藏无 displayDate 也用 jobPostingTime
    if not happen_time:
        happen_time = str(item.get("jobPostingTime", "")) or str(item.get("publishTime", ""))

    return {
        "encrypt_job_id": item.get("number") or item.get("jobId", ""),
        "encrypt_brand_id": item.get("companyNumber") or item.get("companyId", ""),
        "security_id": "",
        "company": item.get("companyName", ""),
        "position": item.get("name") or item.get("positionName", ""),
        "salary": item.get("salary60") or item.get("salary", ""),
        "city": item.get("workCity") or item.get("cityName", ""),
        "district": item.get("cityDistrict", ""),
        "area": item.get("streetName", ""),
        "experience": item.get("workingExp", ""),
        "degree": item.get("education", ""),
        "boss_name": staff.get("staffName", ""),
        "boss_title": staff.get("hrJob", ""),
        "stage": item.get("propertyName") or item.get("property", ""),
        "industry": item.get("industryName", ""),
        "scale": item.get("companySize", ""),
        "action_date": display_date,
        "happen_time": happen_time,
    }


def _extract_items(data: dict, status: str) -> list[dict]:
    d = data.get("data", {})
    # 已投递/收藏: data.data；推荐: data.jobList
    items = d.get("data") or d.get("jobList") or d.get("list") or []
    if isinstance(items, dict):
        items = []
    results = []
    for item in items:
        job = _parse_item(item, status)
        if job["company"] or job["position"]:
            job["status"] = status
            results.append(job)
    return results


def _save_to_storage(jobs: list[dict], today_only: bool = True) -> int:
    from storage import add_application, _load_jobs

    existing = _load_jobs()
    seen = {(j.get("encrypt_job_id", ""), j.get("status", "")) for j in existing}
    passed = 0  # 通过 today_only 的总数（用于显示）
    today = date.today().isoformat()

    for job in jobs:
        jid = job.get("encrypt_job_id", "")
        status = job.get("status", "")
        if not jid:
            continue

        # 推荐类：平台返回的就是今日推荐，直接以当天记录
        is_recommend = "推荐" in status
        if is_recommend:
            record_date = today
        else:
            ts = job.get("happen_time", "")
            if not ts and today_only:
                continue
            record_date = today
            if ts:
                try:
                    from datetime import datetime as _dt
                    record_date = _dt.fromtimestamp(int(ts) / 1000).strftime("%Y-%m-%d")
                except:
                    pass

            if today_only and record_date != today:
                continue

        passed += 1

        key = (jid, status)
        if key in seen:
            continue  # 去重：已存在的不重复存储
        seen.add(key)

        notes_parts = []
        if job.get("salary"):
            notes_parts.append(job["salary"])
        if job.get("city"):
            notes_parts.append(job["city"])

        add_application(
            company=job.get("company", ""),
            position=job.get("position", ""),
            date_str=today,
            status=job.get("status", "已投递"),
            platform="智联招聘",
            notes=", ".join(notes_parts) if notes_parts else "",
            encrypt_job_id=job.get("encrypt_job_id", ""),
            encrypt_brand_id=job.get("encrypt_brand_id", ""),
            city=job.get("city", ""),
            salary=job.get("salary", ""),
            experience=job.get("experience", ""),
            degree=job.get("degree", ""),
            industry=job.get("industry", ""),
            scale=job.get("scale", ""),
            happen_time=job.get("happen_time", ""),
        )
        time.sleep(0.05)

    return passed  # 返回今日过滤后的总数（非去重后的新增数）


def fetch_zhaopin_applied() -> list[dict]:
    cfg = TAB_CONFIG["智联-已投递"]
    data = _request_api(cfg["endpoint"], {**cfg["params"], "index": 1, "pageSize": 50})
    return _extract_items(data, "智联-已投递")


def fetch_zhaopin_collect() -> list[dict]:
    cfg = TAB_CONFIG["智联-收藏"]
    api_base = cfg.get("api_base")
    if api_base:
        data = _request_raw(api_base, {"page": 1, "size": 50, "platform": 13, "version": "0.0.0"})
        items = data.get("data", [])
    else:
        data = _request_api(cfg["endpoint"], {**cfg["params"], "index": 1, "pageSize": 50})
        d = data.get("data", {})
        items = d.get("data") or []
    return _extract_collect_items(items)


def _request_raw(url: str, extra_params: dict = None) -> dict:
    headers, cookies = _get_headers()
    at_token = cookies.get("at", "")
    rt_token = cookies.get("rt", "")
    if not at_token:
        raise PermissionError("请在 users.json 中配置 zhaopin.cookie_json")
    params = {"at": at_token, "rt": rt_token}
    if extra_params:
        params.update(extra_params)
    resp = requests.get(url, params=params, headers=headers, timeout=20)
    resp.raise_for_status()
    return resp.json()


def _extract_collect_items(items: list) -> list[dict]:
    results = []
    for item in items:
        results.append({
            "encrypt_job_id": item.get("number", ""),
            "encrypt_brand_id": item.get("companyNumber", ""),
            "security_id": "",
            "company": item.get("companyName", ""),
            "position": item.get("name", ""),
            "salary": item.get("salary60", ""),
            "city": item.get("workCity", ""),
            "district": item.get("cityDistrict", ""),
            "area": "",
            "experience": item.get("workingExp", ""),
            "degree": item.get("education", ""),
            "boss_name": "",
            "boss_title": "",
            "stage": item.get("property", ""),
            "industry": "",
            "scale": item.get("companySize", ""),
            "action_date": item.get("collectTime", ""),
            "happen_time": str(item.get("collectTimeTimestamp", "")),
            "status": "智联-收藏",
        })
    return results


def fetch_zhaopin_recommend() -> list[dict]:
    data = _request_api("/position/recommend", {"pageSize": 50})
    return _extract_items(data, "智联-推荐")


def fetch_zhaopin_all_recommend() -> list[dict]:
    """翻页获取全部推荐"""
    all_items = []
    page = 1
    while True:
        data = _request_api("/position/recommend", {"pageSize": 50, "page": page})
        items = data.get("data", {}).get("jobList", [])
        all_items.extend(items)
        # 智联推荐没有 hasMore，按返回数量判断
        if len(items) < 40:
            break
        page += 1
        time.sleep(0.3)
    return _extract_items_by_list(all_items, "智联-推荐")


def _extract_items_by_list(items: list, status: str) -> list[dict]:
    results = []
    for item in items:
        job = _parse_item(item, status)
        if job["company"] or job["position"]:
            results.append(job)
    return results


def sync_zhaopin() -> dict:
    results = {"智联-已投递": [], "智联-收藏": [], "智联-推荐": []}
    errors = []
    total_new = 0

    fetchers = {
        "智联-已投递": fetch_zhaopin_applied,
        "智联-收藏": fetch_zhaopin_collect,
        "智联-推荐": fetch_zhaopin_recommend,
    }

    for tab, fetcher in fetchers.items():
        try:
            jobs = fetcher()
            results[tab] = jobs
            n = _save_to_storage(jobs)
            total_new += n
        except PermissionError as e:
            errors.append(f"{tab}: {e}")
            break
        except Exception as e:
            errors.append(f"{tab}: 获取失败 - {e}")

    return {"new": total_new, "counts": {tab: len(jobs) for tab, jobs in results.items()}, "errors": errors}


def export_zhaopin_excel(date_filter: str = None) -> str:
    """导出智联全部（投递+推荐）"""
    r1 = export_zhaopin_delivery_excel(date_filter)
    r2 = export_zhaopin_recommend_excel(date_filter)
    return r1 + "\n" + r2


def export_zhaopin_delivery_excel(date_filter: str = None) -> str:
    """导出智联投递（已投递+收藏）"""
    from storage import _load_jobs, ZHAOPIN_DIR as _report_dir
    from datetime import date as _date
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    if date_filter:
        today = date_filter
    else:
        today = _date.today().isoformat()

    jobs = _load_jobs()
    jobs = [j for j in jobs if j.get("platform") == "智联招聘" and j.get("date") == today and j.get("status") != "智联-推荐"]
    if not jobs:
        return f"{today} 暂无智联投递/收藏记录。"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "智联投递"
    hfont = Font(bold=True, size=11)
    hfill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    bdr = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    link_font = Font(color="0563C1", underline="single")
    icons = {"智联-已投递": "📤 已投递", "智联-收藏": "⭐ 收藏"}

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value="智联招聘 — 投递与收藏").font = Font(bold=True, size=13, color="1F4E79")
    row += 1

    for col, h in enumerate(["序号", "类型", "招聘人", "企业", "企业规模", "招聘职位", "薪资", "要求"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = hfont; c.fill = hfill; c.border = bdr; c.alignment = Alignment(horizontal="center")
    row += 1

    for i, j in enumerate(jobs):
        vals = [i + 1, icons.get(j.get("status", ""), j.get("status", "")), j.get("boss_name", ""), j.get("company", ""),
                " · ".join(p for p in [j.get("industry", ""), j.get("scale", ""), j.get("stage", "")] if p),
                j.get("position", ""), j.get("salary", ""),
                " · ".join(p for p in [j.get("experience", ""), j.get("degree", "")] if p)]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = bdr; c.alignment = Alignment(vertical="center")
            if col == 4:
                cid = j.get("encrypt_brand_id", "")
                if cid: c.font = link_font; c.hyperlink = f"https://jobs.zhaopin.com/{cid}.htm"
            if col == 6:
                jid = j.get("encrypt_job_id", "")
                if jid: c.font = link_font; c.hyperlink = f"https://jobs.zhaopin.com/{jid}.htm"
        row += 1

    for i, w in enumerate([6, 12, 16, 22, 24, 28, 14, 16], 1):
        ws.column_dimensions[chr(64 + i)].width = w

    basename = f"智联招聘_投递_{today}"
    filepath = os.path.join(_report_dir, f"{basename}.xlsx")
    counter = 1
    while os.path.exists(filepath):
        filepath = os.path.join(_report_dir, f"{basename}({counter}).xlsx")
        counter += 1
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    return f"✅ 智联投递 Excel 已生成：`{filepath}`（共 {len(jobs)} 条）"


def export_zhaopin_recommend_excel(date_filter: str = None) -> str:
    """导出智联推荐"""
    from storage import _load_jobs, ZHAOPIN_DIR as _report_dir
    from datetime import date as _date
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    if date_filter:
        today = date_filter
    else:
        today = _date.today().isoformat()

    jobs = _load_jobs()
    jobs = [j for j in jobs if j.get("platform") == "智联招聘" and j.get("status") == "智联-推荐"]
    if not jobs:
        return f"暂无智联推荐数据。"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "智联推荐"
    hfont = Font(bold=True, size=11)
    hfill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    bdr = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    link_font = Font(color="0563C1", underline="single")

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    ws.cell(row=row, column=1, value="智联招聘 — 职位推荐").font = Font(bold=True, size=13, color="1F4E79")
    row += 1

    for col, h in enumerate(["序号", "类型", "企业", "招聘职位", "薪资", "城市", "经验", "学历", "规模"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = hfont; c.fill = hfill; c.border = bdr; c.alignment = Alignment(horizontal="center")
    row += 1

    for i, j in enumerate(jobs):
        vals = [i + 1, "📋 推荐", j.get("company", ""), j.get("position", ""), j.get("salary", ""),
                j.get("city", ""), j.get("experience", ""), j.get("degree", ""), j.get("scale", "")]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = bdr; c.alignment = Alignment(vertical="center")
        row += 1

    for i, w in enumerate([6, 10, 22, 30, 14, 10, 10, 8, 20], 1):
        ws.column_dimensions[chr(64 + i)].width = w

    basename = f"智联招聘_推荐_{today}"
    filepath = os.path.join(_report_dir, f"{basename}.xlsx")
    counter = 1
    while os.path.exists(filepath):
        filepath = os.path.join(_report_dir, f"{basename}({counter}).xlsx")
        counter += 1
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    return f"✅ 智联推荐 Excel 已生成：`{filepath}`（共 {len(jobs)} 条）"


def export_zhaopin_excel_old(date_filter: str = None) -> str:
    """导出智联投递记录为 Excel（已投递+收藏 上部，推荐 下部）"""
    from storage import _load_jobs, ZHAOPIN_DIR as _report_dir
    from datetime import date as _date
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    if date_filter:
        today = date_filter
    else:
        today = _date.today().isoformat()

    jobs = _load_jobs()
    jobs = [j for j in jobs if j.get("platform") == "智联招聘" and j.get("date") == today]
    if not jobs:
        return f"{today} 暂无智联招聘数据。请先说「同步智联」。"

    # 拆分：推荐 vs 已投递+收藏
    recommends = [j for j in jobs if j.get("status") == "智联-推荐"]
    others = [j for j in jobs if j.get("status") != "智联-推荐"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "智联招聘"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_font = Font(bold=True, size=13, color="1F4E79")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    link_font = Font(color="0563C1", underline="single")
    status_icon = {"智联-已投递": "📤 已投递", "智联-收藏": "⭐ 收藏"}

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value="智联招聘 — 投递与收藏")
    cell.font = section_font
    row += 1

    # 上部：已投递 + 收藏
    headers = ["序号", "类型", "招聘人", "企业", "企业规模", "招聘职位", "薪资", "要求"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    row += 1

    for i, j in enumerate(others):
        cid = j.get("encrypt_brand_id", "")
        jid = j.get("encrypt_job_id", "")
        cp_link = f"https://jobs.zhaopin.com/{cid}.htm" if cid else ""
        job_link = j.get("positionURL") or (f"https://jobs.zhaopin.com/{jid}.htm" if jid else "")

        scale_info = " · ".join(p for p in [j.get("industry", ""), j.get("scale", ""), j.get("stage", "")] if p)
        requirements = " · ".join(p for p in [j.get("experience", ""), j.get("degree", "")] if p)
        boss = j.get("boss_name", "")

        values = [i + 1, status_icon.get(j.get("status", ""), j.get("status", "")), boss, j.get("company", ""), scale_info, j.get("position", ""), j.get("salary", ""), requirements]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if col == 4 and cp_link:
                cell.font = link_font
                cell.hyperlink = cp_link
            if col == 6 and job_link:
                cell.font = link_font
                cell.hyperlink = job_link
        row += 1

    if not others:
        ws.cell(row=row, column=1, value="暂无投递/收藏记录")
        row += 1

    row += 1

    # 下部：推荐
    if recommends:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        cell = ws.cell(row=row, column=1, value="智联招聘 — 职位推荐")
        cell.font = section_font
        row += 1

        rec_headers = ["序号", "类型", "企业", "招聘职位", "薪资", "城市", "经验", "学历", "规模"]
        for col, h in enumerate(rec_headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        row += 1

        for i, j in enumerate(recommends):
            cid = j.get("encrypt_brand_id", "")
            jid = j.get("encrypt_job_id", "")
            cp_link = f"https://jobs.zhaopin.com/{cid}.htm" if cid else ""
            job_link = j.get("positionURL") or (f"https://jobs.zhaopin.com/{jid}.htm" if jid else "")

            values = [i + 1, "📋 推荐", j.get("company", ""), j.get("position", ""), j.get("salary", ""), j.get("city", ""), j.get("experience", ""), j.get("degree", ""), j.get("scale", "")]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
                if col == 3 and cp_link:
                    cell.font = link_font
                    cell.hyperlink = cp_link
                if col == 4 and job_link:
                    cell.font = link_font
                    cell.hyperlink = job_link
            row += 1

    # 列宽
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 30

    import os
    basename = f"智联招聘_{today}"
    ext = ".xlsx"
    filepath = os.path.join(_report_dir, f"{basename}{ext}")
    counter = 1
    while os.path.exists(filepath):
        filepath = os.path.join(_report_dir, f"{basename}({counter}){ext}")
        counter += 1

    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    return f"✅ 智联 Excel 已生成：`{filepath}`（共 {len(others) + len(recommends)} 条：投递/收藏 {len(others)} 条 + 推荐 {len(recommends)} 条）"
