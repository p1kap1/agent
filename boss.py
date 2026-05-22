"""BOSS 直聘数据抓取模块

使用前请在 users.json 中配置当前用户的 boss_cookie：
  Chrome: F12 → Application → Cookies → www.zhipin.com → 逐个复制 name 和 value
  然后调用 convert_cookie_json_to_string() 或直接粘贴 cookie 字符串到 boss_cookie 字段
"""

import json
import os
import time
from datetime import date, datetime
from typing import Optional

import requests

from config import DATA_DIR

USERS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "users.json")
# 个人交互数据 API
BOSS_INTERACTION_API = "https://www.zhipin.com/wapi/zprelation/interaction/geekGetJob"
# 已投递 API
BOSS_DELIVERY_API = "https://www.zhipin.com/wapi/zprelation/resume/geekDeliverList"
# 面试 API  
BOSS_INTERVIEW_API = "https://www.zhipin.com/wapi/zpinterview/geek/interview/list"
# 每日推荐 API（优先用精准推荐，失败回退到通用推荐）
BOSS_RECOMMEND_DAILY = "https://www.zhipin.com/wapi/zpgeek/recommend/job/list.json"

TAB_PARAMS = {
    "沟通过": {"api": "interaction", "tag": 5, "isActive": "true"},
    "已投递": {"api": "delivery"},
    "面试":   {"api": "interview"},
    "感兴趣": {"api": "interaction", "tag": 4, "isActive": "true"},
    "每日推荐": {"api": "daily_recommend"},
}
DEFAULT_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/125.0.0.0 Safari/537.36"
)


def convert_cookie_json_to_string(cookie_json: list[dict]) -> str:
    """将浏览器导出的 JSON 格式 Cookie 转为请求头所需的字符串格式

    输入: [{"name": "zp_at", "value": "xxx"}, ...]
    输出: "zp_at=xxx; wt2=yyy; ..."
    """
    pairs = []
    for item in cookie_json:
        name = item.get("name", "")
        value = item.get("value", "")
        if name:
            pairs.append(f"{name}={value}")
    return "; ".join(pairs)


def _load_users() -> dict:
    if not os.path.exists(USERS_FILE):
        return {"active_user": "default", "users": {}}
    with open(USERS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def _save_users(data: dict):
    with open(USERS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def get_active_user_config() -> dict:
    """获取当前活跃用户的配置字典"""
    data = _load_users()
    active = data.get("active_user", "default")
    users = data.get("users", {})
    return users.get(active, {})


def set_active_user(username: str):
    """切换活跃用户"""
    data = _load_users()
    if username not in data.get("users", {}):
        raise ValueError(f"用户 '{username}' 不存在于 users.json 中")
    data["active_user"] = username
    _save_users(data)


def set_user_cookie(username: str, cookie_str: str):
    """更新指定用户的 Boss直聘 Cookie"""
    data = _load_users()
    if username not in data.get("users", {}):
        data.setdefault("users", {})[username] = {}
    data["users"][username]["boss_cookie"] = cookie_str.strip()
    _save_users(data)


def _get_cookie() -> str:
    cfg = get_active_user_config()
    return cfg.get("boss_cookie", "")


def _make_headers() -> dict:
    return {
        "User-Agent": DEFAULT_UA,
        "Cookie": _get_cookie(),
        "Referer": "https://www.zhipin.com/web/geek/recommend",
        "Accept": "application/json, text/plain, */*",
    }


def _request_api(url: str, params: dict = None) -> dict:
    """调用 Boss直聘 API，返回解析后的 JSON"""
    cookie = _get_cookie()
    if not cookie:
        raise PermissionError("请先在 users.json 中配置 boss_cookie")

    resp = requests.get(
        url,
        headers=_make_headers(),
        params=params or {},
        timeout=20,
    )
    if resp.status_code == 401 or resp.status_code == 302:
        raise PermissionError("Cookie 已过期或无效，请更新 boss_cookie")

    resp.raise_for_status()
    data = resp.json()

    code = data.get("code")
    if code is not None and code != 0:
        msg = data.get("message", "未知错误")
        raise RuntimeError(f"BOSS API 返回错误 (code={code}): {msg}")

    return data


def _parse_job_item(item: dict) -> dict:
    """将 API 返回的单条职位数据解析为统一格式"""
    return {
        "encrypt_job_id": item.get("encryptJobId", ""),
        "encrypt_brand_id": item.get("encryptBrandId", ""),
        "security_id": item.get("securityId", ""),
        "job_id": (
            item.get("encryptJobId")
            or item.get("encryptId")
            or item.get("encryptBossId")
            or ""
        ),
        "company": item.get("brandName", ""),
        "position": item.get("jobName", ""),
        "salary": item.get("salaryDesc") or item.get("jobSalary", ""),
        "city": item.get("cityName") or item.get("city", ""),
        "district": item.get("areaDistrict", ""),
        "area": item.get("businessDistrict", ""),
        "experience": item.get("jobExperience", ""),
        "degree": item.get("jobDegree", ""),
        "boss_name": item.get("bossName", ""),
        "boss_title": item.get("bossTitle", ""),
        "stage": item.get("brandStageName", ""),
        "industry": item.get("industryName") or item.get("brandIndustry", ""),
        "scale": item.get("scaleName") or item.get("brandScaleName", ""),
        "action_date": item.get("actionDateDesc", ""),
        "happen_time": item.get("happenTime", ""),
    }


def _parse_interview_item(item: dict) -> dict:
    """将面试接口返回的数据解析为统一格式"""
    return {
        "encrypt_job_id": item.get("encryptJobId") or item.get("encryptId", ""),
        "encrypt_brand_id": item.get("encryptBrandId", ""),
        "security_id": item.get("securityId", ""),
        "job_id": item.get("encryptJobId") or item.get("encryptId", ""),
        "company": item.get("brandName", ""),
        "position": item.get("jobName", ""),
        "salary": item.get("salaryDesc") or item.get("jobSalary", ""),
        "city": item.get("cityName", "") or item.get("city", ""),
        "district": item.get("areaDistrict", ""),
        "area": item.get("businessDistrict", ""),
        "experience": item.get("jobExperience", ""),
        "degree": item.get("jobDegree", ""),
        "boss_name": item.get("bossName", ""),
        "boss_title": item.get("bossTitle", ""),
        "stage": item.get("brandStageName", ""),
        "industry": item.get("industryName") or item.get("brandIndustry", ""),
        "scale": item.get("scaleName") or item.get("brandScaleName", ""),
        "action_date": item.get("actionDateDesc", ""),
        # 面试特有字段
        "interview_time": item.get("interviewTime", ""),
        "interview_addr": item.get("interviewAddr", ""),
        "interview_status": item.get("interviewStatus", ""),
        "video_link": item.get("videoLink", ""),
        "happen_time": item.get("happenTime", "") or item.get("appointmentDate", ""),
    }


def _extract_jobs_from_items(items: list, status: str, api_type: str = "interaction") -> list[dict]:
    """将原始 item 列表解析为统一格式"""
    results = []
    for item in items:
        if api_type == "interview":
            job = _parse_interview_item(item)
        else:
            job = _parse_job_item(item)
        if job.get("company") or job.get("position"):
            job["status"] = status
            results.append(job)
    return results


def _load_existing_ids() -> set:
    """从 applications.json 读取已有的 encrypt_job_id 集合，用于去重"""
    from storage import _load_jobs
    jobs = _load_jobs()
    return {j.get("encrypt_job_id", "") for j in jobs} | {j.get("job_id", "") for j in jobs}


def _fetch_tab(tab: str) -> list[dict]:
    """统一获取指定 tab 的数据（自动翻页获取全部）"""
    cfg = TAB_PARAMS.get(tab, {})
    api_type = cfg.get("api", "interaction")
    page = 1
    page_size = 50  # 服务端限制最大 15

    all_items = []

    while True:
        if api_type == "delivery":
            url = BOSS_DELIVERY_API
            params = {"page": page, "pageSize": page_size}
        elif api_type == "interview":
            url = BOSS_INTERVIEW_API
            params = {"page": page, "pageSize": page_size}
        elif api_type == "daily_recommend":
            url = BOSS_RECOMMEND_DAILY
            params = {"tab": 0, "page": page, "pageSize": page_size}
        else:
            url = BOSS_INTERACTION_API
            params = {"tag": cfg.get("tag"), "page": page, "pageSize": page_size}
            if "isActive" in cfg:
                params["isActive"] = cfg["isActive"]

        data = _request_api(url, params)
        zp_data = data.get("zpData", {})

        if api_type == "interview":
            items = zp_data.get("interviewList", [])
            has_more = zp_data.get("hasMore", False)
        else:
            items = zp_data.get("cardList") or zp_data.get("jobList") or []
            has_more = zp_data.get("hasMore", False)

        all_items.extend(items)

        # 每日推荐 hasMore 不可靠，按页数上限翻页；其他按 hasMore
        if api_type == "daily_recommend":
            if page >= 3:
                break
        elif not has_more:
            break
        page += 1
        time.sleep(1.0 if api_type == "daily_recommend" else 0.3)

    return _extract_jobs_from_items(all_items, tab, api_type)


def fetch_boss_channels() -> list[dict]:
    """获取「沟通过」列表"""
    return _fetch_tab("沟通过")


def fetch_boss_applied() -> list[dict]:
    """获取「已投递」列表"""
    return _fetch_tab("已投递")


def fetch_boss_interviews() -> list[dict]:
    """获取「面试」列表"""
    return _fetch_tab("面试")


def fetch_boss_interested() -> list[dict]:
    """获取「感兴趣」列表"""
    return _fetch_tab("感兴趣")


def fetch_daily_recommend() -> list[dict]:
    """获取「每日推荐」列表"""
    return _fetch_tab("每日推荐")


def _save_jobs_to_storage(jobs: list[dict], today_only: bool = True) -> int:
    """将职位列表写入 applications.json，同批次同状态去重。today_only 只存今天的数据"""
    from storage import add_application
    import datetime as _dt_local

    seen = set()
    new_count = 0
    today = _dt_local.date.today()

    for job in jobs:
        jid = job.get("encrypt_job_id", "") or job.get("job_id", "")
        if not jid:
            continue
        status = job.get("status", "")
        key = (jid, status)
        if key in seen:
            continue
        seen.add(key)

        # 按 happenTime 计算日期
        ts = job.get("happen_time", "")
        record_date = today.isoformat()
        if ts:
            try:
                record_date = _dt_local.datetime.fromtimestamp(int(ts) / 1000).strftime("%Y-%m-%d")
            except (ValueError, OSError):
                pass

        if today_only and record_date != today.isoformat():
            continue

        notes_parts = []
        if job.get("salary"):
            notes_parts.append(job["salary"])
        if job.get("city"):
            notes_parts.append(job["city"])
        if job.get("stage"):
            notes_parts.append(job["stage"])
        if job.get("industry"):
            notes_parts.append(job["industry"])

        notes_parts = []
        if job.get("salary"):
            notes_parts.append(job["salary"])
        if job.get("city"):
            notes_parts.append(job["city"])
        if job.get("stage"):
            notes_parts.append(job["stage"])
        if job.get("industry"):
            notes_parts.append(job["industry"])

        add_application(
            company=job.get("company", ""),
            position=job.get("position", ""),
            date_str=record_date,
            status=job.get("status", "已投递"),
            platform="Boss直聘",
            notes=", ".join(notes_parts) if notes_parts else "",
            encrypt_job_id=job.get("encrypt_job_id", ""),
            encrypt_brand_id=job.get("encrypt_brand_id", ""),
            security_id=job.get("security_id", ""),
            city=job.get("city", ""),
            district=job.get("district", ""),
            area=job.get("area", ""),
            experience=job.get("experience", ""),
            degree=job.get("degree", ""),
            salary=job.get("salary", ""),
            boss_name=job.get("boss_name", ""),
            boss_title=job.get("boss_title", ""),
            stage=job.get("stage", ""),
            industry=job.get("industry", ""),
            scale=job.get("scale", ""),
            happen_time=job.get("happen_time", ""),
        )
        new_count += 1
        time.sleep(0.05)

    return new_count


def sync_all(today_only: bool = True) -> dict:
    """同步全部四个模块，today_only=True 只保存今天的数据"""
    results = {"沟通过": [], "已投递": [], "面试": [], "感兴趣": []}
    errors = []
    total_new = 0

    fetchers = {
        "沟通过": fetch_boss_channels,
        "已投递": fetch_boss_applied,
        "面试": fetch_boss_interviews,
        "感兴趣": fetch_boss_interested,
    }

    for tab, fetcher in fetchers.items():
        try:
            jobs = fetcher()
            results[tab] = jobs
            n = _save_jobs_to_storage(jobs)
            total_new += n
        except PermissionError as e:
            errors.append(f"{tab}: {e}")
            break
        except Exception as e:
            errors.append(f"{tab}: 获取失败 - {e}")

    return {
        "new": total_new,
        "counts": {tab: len(jobs) for tab, jobs in results.items()},
        "errors": errors,
    }


def boss_job_summary() -> str:
    """统计汇总四个模块的投递数据"""
    from storage import _load_jobs

    jobs = _load_jobs()
    if not jobs:
        return "暂无投递记录。请先说「同步投递」拉取数据。"

    by_status = {}
    for j in jobs:
        by_status.setdefault(j["status"], [])
        by_status[j["status"]].append(j)

    total = len(jobs)
    lines = [
        f"# 📊 Boss直聘投递统计",
        f"",
        f"**总计**: {total} 家公司",
        f"",
    ]

    for status in ["沟通过", "已投递", "面试", "感兴趣", "不合适"]:
        items = by_status.get(status, [])
        lines.append(f"- **{status}**: {len(items)} 家")

    lines.append("")
    lines.append("## 详细列表")
    lines.append("")

    for status in ["面试", "感兴趣", "已投递", "沟通过", "不合适"]:
        items = by_status.get(status, [])
        if not items:
            continue
        lines.append(f"### {status}（{len(items)}）")
        for j in items:
            line = f"- **{j['company']}** — {j['position']}"
            if j.get("notes"):
                line += f"（{j['notes']}）"
            line += f" `{j['date']}`"
            lines.append(line)
        lines.append("")

    return "\n".join(lines)


def export_excel(status_filter: str = None, date_filter: str = None) -> str:
    """导出投递记录为 Excel 文件（默认只导出今天的数据，面试单独区域）"""
    from storage import _load_jobs
    from datetime import date as _date
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    if date_filter:
        today = date_filter
    else:
        today = _date.today().isoformat()

    jobs = _load_jobs()
    jobs = [j for j in jobs if j.get("date") == today]
    if status_filter:
        jobs = [j for j in jobs if j.get("status") == status_filter]
    # 排除每日推荐和智联
    jobs = [j for j in jobs if j.get("platform") == "Boss直聘" and j.get("status") != "每日推荐" and j.get("status") != "智联-推荐"]
    if not jobs:
        return f"{today} 暂无投递记录可导出。请先说「同步投递」拉取数据。"

    # 拆分：面试和其他
    interviews = [j for j in jobs if j.get("status") == "面试"]
    others = [j for j in jobs if j.get("status") != "面试"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Boss直聘投递记录"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_font = Font(bold=True, size=13, color="1F4E79")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    link_font = Font(color="0563C1", underline="single")
    status_icon = {"沟通过": "💬 沟通过", "已投递": "📤 已投递", "感兴趣": "⭐ 感兴趣", "面试": "🎯 面试", "不合适": "❌ 不合适"}

    # ===== 第一部分：沟通过 + 已投递 + 感兴趣 =====
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    cell = ws.cell(row=row, column=1, value="Boss直聘投递记录")
    cell.font = section_font
    row += 1

    main_headers = ["序号", "类型", "招聘人", "企业", "企业规模", "招聘职位", "薪资", "要求"]
    for col, h in enumerate(main_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    row += 1

    for i, j in enumerate(others):
        ebid = j.get("encrypt_brand_id", "")
        ejid = j.get("encrypt_job_id", "")
        sid = j.get("security_id", "")
        company_link = f"https://www.zhipin.com/gongsi/{ebid}.html?ka=personal_added_brand_{ebid}" if ebid else ""
        scale_info = "，".join(p for p in [j.get("industry", ""), j.get("stage", ""), j.get("scale", "")] if p)
        requirements = "，".join(p for p in [j.get("experience", ""), j.get("degree", "")] if p)
        boss = f"{j.get('boss_name', '')}{j.get('boss_title', '')}"
        status_label = status_icon.get(j.get("status", ""), j.get("status", ""))
        values = [i + 1, status_label, boss, j.get("company", ""), scale_info, j.get("position", ""), j.get("salary", ""), requirements]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if col == 4 and company_link:  # 企业名可点击
                cell.font = link_font
                cell.hyperlink = company_link
            if col == 6 and ejid and sid:  # 招聘职位可点击
                job_link = f"https://www.zhipin.com/job_detail/{ejid}.html?securityId={sid}&ka=personal_added_job_{ejid}"
                cell.font = link_font
                cell.hyperlink = job_link
        row += 1

    row += 1  # 空一行

    # ===== 第二部分：面试（单独格式）=====
    if interviews:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        cell = ws.cell(row=row, column=1, value="面试记录")
        cell.font = section_font
        row += 1

        interview_headers = ["序号", "类型", "企业", "招聘职位", "薪资", "面试时间", "面试地址", "状态", "视频链接"]
        for col, h in enumerate(interview_headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        row += 1

        for i, j in enumerate(interviews):
            ebid = j.get("encrypt_brand_id", "")
            ejid = j.get("encrypt_job_id", "")
            sid = j.get("security_id", "")
            company_link = f"https://www.zhipin.com/gongsi/{ebid}.html?ka=personal_added_brand_{ebid}" if ebid else ""
            job_link_val = ""
            if ejid and sid:
                job_link_val = f"https://www.zhipin.com/job_detail/{ejid}.html?securityId={sid}&ka=personal_added_job_{ejid}"

            values = [
                i + 1,
                "🎯 面试",
                j.get("company", ""),
                j.get("position", ""),
                j.get("salary", ""),
                j.get("interview_time", ""),
                j.get("interview_addr", ""),
                j.get("interview_status", ""),
                j.get("video_link", ""),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
                if col == 3 and company_link:
                    cell.font = link_font
                    cell.hyperlink = company_link
                if col == 4 and job_link_val:
                    cell.font = link_font
                    cell.hyperlink = job_link_val
                if col == 9 and j.get("video_link"):
                    cell.font = link_font
                    cell.hyperlink = j["video_link"]
            row += 1

    # 列宽
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 28
    ws.column_dimensions["H"].width = 30
    ws.column_dimensions["I"].width = 40

    # 文件名
    import os
    from storage import BOSS_DIR as _report_dir
    basename = f"Boss直聘_投递_{today}"
    ext = ".xlsx"
    filepath = os.path.join(_report_dir, f"{basename}{ext}")
    counter = 1
    while os.path.exists(filepath):
        filepath = os.path.join(_report_dir, f"{basename}({counter}){ext}")
        counter += 1

    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    total = len(others) + len(interviews)
    return f"✅ Excel 已生成：`{filepath}`（共 {total} 条：沟通过/已投递/感兴趣 {len(others)} 条 + 面试 {len(interviews)} 条）"


def export_all_reports(date_filter: str = None) -> str:
    """导出全部数据（Boss + 智联 + 猎聘）"""
    results = []
    results.append(export_excel(date_filter=date_filter))
    results.append(export_daily_recommend_excel(date_filter=date_filter))
    try:
        import zhaopin
        results.append(zhaopin.export_zhaopin_excel(date_filter=date_filter))
    except Exception:
        pass
    try:
        import liepin
        results.append(liepin.export_liepin_excel(date_filter=date_filter))
    except Exception:
        pass
    return "\n".join(results)


def export_daily_recommend_excel(date_filter: str = None) -> str:
    """单独导出每日推荐为 Excel"""
    from storage import _load_jobs, BOSS_DIR as _report_dir
    from datetime import date as _date
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    if date_filter:
        today = date_filter
    else:
        today = _date.today().isoformat()

    jobs = _load_jobs()
    jobs = [j for j in jobs if j.get("status") == "每日推荐" and j.get("date") == today]
    if not jobs:
        return f"{today} 暂无每日推荐数据。请先说「同步每日推荐」。"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "每日推荐"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    link_font = Font(color="0563C1", underline="single")

    headers = ["序号", "公司", "职位", "薪资", "城市", "经验", "学历", "规模", "行业"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for i, j in enumerate(jobs, 2):
        values = [
            i - 1,
            j.get("company", ""),
            j.get("position", ""),
            j.get("salary", ""),
            j.get("city", ""),
            j.get("experience", ""),
            j.get("degree", ""),
            j.get("scale", ""),
            j.get("industry", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border
            if col == 2:  # 企业名可点击
                ebid = j.get("encrypt_brand_id", "")
                if ebid:
                    cell.font = link_font
                    cell.hyperlink = f"https://www.zhipin.com/gongsi/{ebid}.html"

    widths = [6, 26, 40, 16, 10, 10, 8, 16, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    import os
    basename = f"Boss直聘_每日推荐_{today}"
    ext = ".xlsx"
    filepath = os.path.join(_report_dir, f"{basename}{ext}")
    counter = 1
    while os.path.exists(filepath):
        filepath = os.path.join(_report_dir, f"{basename}({counter}){ext}")
        counter += 1

    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    return f"✅ 每日推荐 Excel 已生成：`{filepath}`（共 {len(jobs)} 条）"


def list_exports() -> str:
    """列出所有已导出的 Excel 文件"""
    import os
    from storage import BOSS_DIR as _report_dir
    if not os.path.exists(_report_dir):
        return "尚未导出过 Excel 文件。"
    files = sorted(
        [f for f in os.listdir(_report_dir) if f.startswith("Boss直聘_") and f.endswith(".xlsx")],
        reverse=True,
    )
    if not files:
        return "尚未导出过 Excel 文件。"
    lines = [f"共 {len(files)} 个导出文件：", ""]
    for f in files[:10]:
        path = os.path.join(_report_dir, f)
        size = os.path.getsize(path)
        lines.append(f"- `{f}` ({size} bytes)")
    return "\n".join(lines)


def show_daily_recommend_table() -> str:
    """展示每日推荐岗位列表（最多10条）"""
    from storage import _load_jobs
    from datetime import date as _date

    today = _date.today().isoformat()
    jobs = _load_jobs()
    jobs = [j for j in jobs if j.get("status") == "每日推荐" and j.get("date") == today]

    if not jobs:
        return f"今天暂无每日推荐数据。请先说「同步每日推荐」。"

    display = jobs[:10]
    lines = [f"每日推荐岗位（{today}，共 {len(jobs)} 条）：", ""]
    for j in display:
        company = j.get("company", "")
        position = j.get("position", "")
        salary = j.get("salary", "")
        city = j.get("city", "")
        parts = [company, position]
        if salary:
            parts.append(salary)
        if city:
            parts.append(city)
        lines.append(f"- {' · '.join(parts)}")

    if len(jobs) > 10:
        lines.append(f"\n（共 {len(jobs)} 条，展示前10条。说「导出Excel」获取完整列表）")

    return "\n".join(lines)


def show_application_table(status: str = None, date_str: str = None) -> str:
    """展示投递岗位列表（最多10条）"""
    from storage import _load_jobs
    from datetime import date as _date

    if not date_str:
        date_str = _date.today().isoformat()

    jobs = _load_jobs()
    jobs = [j for j in jobs if j.get("date") == date_str]
    if status:
        jobs = [j for j in jobs if j.get("status") == status]
    jobs = [j for j in jobs if j.get("status") != "每日推荐"]

    if not jobs:
        return f"{date_str} 暂无投递记录。"

    status_icon = {"沟通过": "💬", "已投递": "📤", "面试": "🎯", "感兴趣": "⭐", "不合适": "❌"}
    display = jobs[:10]
    lines = [f"投递岗位（{date_str}，共 {len(jobs)} 条）：", ""]
    for j in display:
        st = status_icon.get(j.get("status", ""), "")
        company = j.get("company", "")
        position = j.get("position", "")
        salary = j.get("salary", "")
        city = j.get("city", "")
        parts = [st, company, position]
        if salary:
            parts.append(salary)
        if city:
            parts.append(city)
        lines.append(f"- {' · '.join(parts)}")

    if len(jobs) > 10:
        lines.append(f"\n（共 {len(jobs)} 条，展示前10条。说「导出Excel」获取完整列表）")

    return "\n".join(lines)


if __name__ == "__main__":
    result = sync_all()
    print(f"同步完成: 新增 {result['new']} 条")
    for tab, count in result["counts"].items():
        print(f"  {tab}: {count} 条")
    if result["errors"]:
        for e in result["errors"]:
            print(f"  ⚠ {e}")

    print()
    print(boss_job_summary())
