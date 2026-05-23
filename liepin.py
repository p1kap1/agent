"""猎聘数据抓取模块

API 端点 (全部 POST):
- 已投递/已查看/面试: com.liepin.capply.pc.apply-list
- 我的收藏: com.liepin.c.job-favorite.get-favorite-job-page
- 推荐: com.liepin.csearch.home-recommend-job-new

认证: lt_auth cookie + XSRF-TOKEN + X-Fscp-* 头 + data 包裹
"""

import json
import os
import time
import uuid
from datetime import date, datetime

import requests

from config import DATA_DIR

LIPIN_BASE = "https://api-c.liepin.com/api"

TAB_CONFIG = {
    "猎聘-已投递": {"endpoint": "com.liepin.capply.pc.apply-list", "body": {"data": {"status": "applied", "pageSize": 50, "page": 1}}},
    "猎聘-已查看": {"endpoint": "com.liepin.capply.pc.apply-list", "body": {"data": {"status": "viewed", "pageSize": 50, "page": 1}}},
    "猎聘-面试":   {"endpoint": "com.liepin.capply.pc.apply-list", "body": {"data": {"status": "interview", "pageSize": 50, "page": 1}}},
    "猎聘-收藏":   {"endpoint": "com.liepin.c.job-favorite.get-favorite-job-page", "body": {"data": {"pageSize": 50, "page": 1}}},
    "猎聘-推荐": {"endpoint": "com.liepin.csearch.home-recommend-job-new", "body": None},
}


def _load_config() -> dict:
    users_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "users.json")
    if not os.path.exists(users_file):
        return {}
    with open(users_file, "r", encoding="utf-8") as f:
        users = json.load(f)
    active = users.get("active_user", "default")
    return users.get("users", {}).get(active, {}).get("liepin", {})


def _get_xsrf() -> str:
    cfg = _load_config()
    cookie_str = cfg.get("cookie", "")
    for pair in cookie_str.split("; "):
        if pair.startswith("XSRF-TOKEN="):
            return pair.split("=", 1)[1]
    return ""


def _get_selected_expect() -> str:
    """动态获取用户期望设置 JSON"""
    data = _request_api("com.liepin.csearch.pc.get-valid-expect-info", {"data": {}})
    expects = data.get("data", {}).get("validExpects", [])
    if expects:
        return json.dumps(expects[0], ensure_ascii=False)
    return ""


def _request_api(endpoint: str, body: dict = None) -> dict:
    cfg = _load_config()
    cookie_str = cfg.get("cookie", "")
    if not cookie_str:
        raise PermissionError("请在 users.json 中配置 liepin.cookie")

    xsrf = "2rLZjSoPRPm1gx_E-mm54A"
    for pair in cookie_str.split("; "):
        if pair.startswith("XSRF-TOKEN="):
            xsrf = pair.split("=", 1)[1]

    hdrs = {
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7",
        "Content-Type": "application/json;charset=UTF-8",
        "Origin": "https://c.liepin.com",
        "Referer": "https://c.liepin.com/",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-site",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36",
        "X-Client-Type": "web",
        "X-Fscp-Bi-Stat": '{"location":"https://c.liepin.com/?time=' + str(int(time.time() * 1000)) + '"}',
        "X-Fscp-Std-Info": '{"client_id":"40106"}',
        "X-Fscp-Trace-Id": str(uuid.uuid4()),
        "X-Fscp-Version": "1.1",
        "X-Requested-With": "XMLHttpRequest",
        "X-XSRF-TOKEN": xsrf,
        "sec-ch-ua": '"Chromium";v="148", "Google Chrome";v="148", "Not/A)Brand";v="99"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
    }

    s = requests.Session()
    req = requests.Request("POST", f"{LIPIN_BASE}/{endpoint}", headers=hdrs, json=body or {})
    prep = s.prepare_request(req)
    prep.headers.update(hdrs)
    prep.headers["Cookie"] = cookie_str

    resp = s.send(prep, timeout=20)
    if resp.status_code in (401, 403):
        raise PermissionError("猎聘登录态已过期，请更新 Cookie")
    resp.raise_for_status()
    data = resp.json()

    flag = data.get("flag")
    if flag is not None and flag != 1 and flag != 0:
        raise RuntimeError(f"猎聘 API 错误 (flag={flag})")
    return data


def _extract_items(data: dict, status: str) -> list[dict]:
    d = data.get("data", {})
    items = d.get("data") or d.get("datas") or d.get("jobCardList") or d.get("list") or []
    results = []
    for item in items:
        # 收藏的 item 结构不同，job/comp/recruiter 在 jobCardDto 里
        card = item.get("jobCardDto") or item
        job = (card.get("job") or item.get("job") or {}).copy() if isinstance(card, dict) else {}
        comp = (card.get("comp") or item.get("comp") or {}).copy() if isinstance(card, dict) else {}
        di = card.get("dataInfo") or item.get("dataInfo") or {}
        recruiter = card.get("recruiter") or item.get("recruiter") or {}
        # 确保是 dict
        if isinstance(job, str): job = {}
        if isinstance(comp, str): comp = {}
        if isinstance(di, str): di = {}
        if isinstance(recruiter, str): recruiter = {}

        results.append({
            "encrypt_job_id": str(job.get("jobId") or item.get("jobId", "")),
            "encrypt_brand_id": str(comp.get("compId") or comp.get("companyId", "")),
            "company": comp.get("compName") or comp.get("name", "") or item.get("company", ""),
            "position": job.get("title") or job.get("jobTitle", "") or item.get("title", ""),
            "salary": job.get("salary") or di.get("salary", "") or item.get("salary", ""),
            "city": job.get("city") or job.get("workCity", "") or di.get("city", ""),
            "district": job.get("district", ""),
            "experience": job.get("requireWorkYears") or di.get("requireWorkYears", "") or item.get("experience", ""),
            "degree": job.get("requireDegree") or di.get("requireDegree", "") or item.get("degree", ""),
            "boss_name": recruiter.get("staffName") or recruiter.get("recruiterName") or recruiter.get("name", "") or item.get("publisherName", ""),
            "boss_title": recruiter.get("hrJob") or recruiter.get("title") or item.get("publisherTitle", ""),
            "stage": comp.get("compStage") or comp.get("stage", "") or item.get("stage", ""),
            "industry": comp.get("compIndustry") or comp.get("industry", "") or di.get("industry", "") or item.get("industry", ""),
            "scale": comp.get("compScale") or comp.get("scale", "") or item.get("scale", ""),
            "action_date": di.get("applyTime") or di.get("time", "") or item.get("favoriteTime", "") or item.get("applyTime", ""),
            "happen_time": str(di.get("applyTimeStamp") or di.get("timestamp", "") or item.get("favoriteTime", "") or item.get("applyTimestamp", "")),
            "status": status,
        })
    return results


def _save_to_storage(jobs: list[dict], today_only: bool = True) -> int:
    from storage import add_application, _load_jobs

    existing = _load_jobs()
    seen = {(j.get("encrypt_job_id", ""), j.get("status", "")) for j in existing}
    passed = 0  # 通过 today_only 的总数（用于显示）
    today = date.today().isoformat()

    for job in jobs:
        jid = job.get("encrypt_job_id", "")
        status = job.get("status", "")
        if not jid:
            continue

        # 推荐类：平台返回的就是今日推荐，直接以当天记录
        is_recommend = "推荐" in status
        if is_recommend:
            record_date = today
        else:
            ts = job.get("happen_time", "")
            if not ts and today_only:
                continue
            record_date = today
            if ts:
                try:
                    from datetime import datetime as _dt
                    record_date = _dt.fromtimestamp(int(ts) / 1000).strftime("%Y-%m-%d")
                except:
                    pass

            if today_only and record_date != today:
                continue

        passed += 1

        key = (jid, status)
        if key in seen:
            continue  # 去重：已存在的不重复存储
        seen.add(key)

        notes_parts = []
        if job.get("salary"):
            notes_parts.append(job["salary"])
        if job.get("city"):
            notes_parts.append(job["city"])

        add_application(
            company=job.get("company", ""),
            position=job.get("position", ""),
            date_str=record_date,
            status=job.get("status", "已投递"),
            platform="猎聘",
            notes=", ".join(notes_parts) if notes_parts else "",
            encrypt_job_id=job.get("encrypt_job_id", ""),
            encrypt_brand_id=job.get("encrypt_brand_id", ""),
            city=job.get("city", ""),
            salary=job.get("salary", ""),
            experience=job.get("experience", ""),
            degree=job.get("degree", ""),
            industry=job.get("industry", ""),
            scale=job.get("scale", ""),
            happen_time=job.get("happen_time", ""),
        )
        time.sleep(0.05)

    return passed  # 返回今日过滤后的总数（非去重后的新增数）


def fetch_liepin_applied() -> list[dict]:
    cfg = TAB_CONFIG["猎聘-已投递"]
    data = _request_api(cfg["endpoint"], cfg.get("body"))
    return _extract_items(data, "猎聘-已投递")


def fetch_liepin_viewed() -> list[dict]:
    cfg = TAB_CONFIG["猎聘-已查看"]
    data = _request_api(cfg["endpoint"], cfg.get("body"))
    return _extract_items(data, "猎聘-已查看")


def fetch_liepin_interview() -> list[dict]:
    cfg = TAB_CONFIG["猎聘-面试"]
    data = _request_api(cfg["endpoint"], cfg.get("body"))
    return _extract_items(data, "猎聘-面试")


def fetch_liepin_collect() -> list[dict]:
    cfg = TAB_CONFIG["猎聘-收藏"]
    data = _request_api(cfg["endpoint"], cfg.get("body"))
    return _extract_items(data, "猎聘-收藏")


def fetch_liepin_recommend() -> list[dict]:
    expect_json = _get_selected_expect()
    body = {"data": {
        "operateKind": "LOGIN", "sortType": "PC_STU_HP_MIX",
        "existFallbackResult": False,
    }}
    if expect_json:
        body["data"]["selectedExpect"] = expect_json
    data = _request_api(TAB_CONFIG["猎聘-推荐"]["endpoint"], body)
    return _extract_items(data, "猎聘-推荐")


def sync_liepin() -> dict:
    results = {}
    errors = []
    total_new = 0

    fetchers = {
        "猎聘-已投递": fetch_liepin_applied,
        "猎聘-已查看": fetch_liepin_viewed,
        "猎聘-面试": fetch_liepin_interview,
        "猎聘-收藏": fetch_liepin_collect,
        "猎聘-推荐": fetch_liepin_recommend,
    }

    for tab, fetcher in fetchers.items():
        try:
            jobs = fetcher()
            results[tab] = jobs
            n = _save_to_storage(jobs)
            total_new += n
        except PermissionError as e:
            errors.append(f"{tab}: {e}")
            break
        except Exception as e:
            errors.append(f"{tab}: 获取失败 - {e}")

    return {"new": total_new, "counts": {tab: len(jobs) for tab, jobs in results.items()}, "errors": errors}


def export_liepin_excel(date_filter: str = None) -> str:
    """导出猎聘全部（投递+推荐）"""
    r1 = export_liepin_delivery_excel(date_filter)
    r2 = export_liepin_recommend_excel(date_filter)
    return r1 + "\n" + r2


def export_liepin_delivery_excel(date_filter: str = None) -> str:
    """只导出猎聘投递（已投递/已查看/面试/收藏）"""
    from storage import _load_jobs
    from datetime import date as _date
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    liepin_dir = os.path.join(DATA_DIR, "reports", "liepin")
    os.makedirs(liepin_dir, exist_ok=True)

    if date_filter:
        today = date_filter
    else:
        today = _date.today().isoformat()

    jobs = _load_jobs()
    jobs = [j for j in jobs if j.get("platform") == "猎聘" and j.get("date") == today and j.get("status") != "猎聘-推荐"]
    if not jobs:
        return f"{today} 暂无猎聘投递记录。"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "猎聘投递"
    hfont = Font(bold=True, size=11)
    hfill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    bdr = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    link_font = Font(color="0563C1", underline="single")
    icons = {"猎聘-已投递": "📤 已投递", "猎聘-已查看": "👁 已查看", "猎聘-面试": "🎯 面试", "猎聘-收藏": "⭐ 收藏"}

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value="猎聘 — 投递记录").font = Font(bold=True, size=13, color="1F4E79")
    row += 1

    for col, h in enumerate(["序号", "类型", "招聘人", "企业", "企业规模", "招聘职位", "薪资", "要求"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = hfont; c.fill = hfill; c.border = bdr; c.alignment = Alignment(horizontal="center")
    row += 1

    for i, j in enumerate(jobs):
        vals = [i + 1, icons.get(j.get("status", ""), j.get("status", "")), j.get("boss_name", ""), j.get("company", ""),
                " · ".join(p for p in [j.get("industry", ""), j.get("scale", ""), j.get("stage", "")] if p),
                j.get("position", ""), j.get("salary", ""),
                " · ".join(p for p in [j.get("experience", ""), j.get("degree", "")] if p)]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = bdr; c.alignment = Alignment(vertical="center")
            if col == 4:
                cid = j.get("encrypt_brand_id", "")
                if cid: c.font = link_font; c.hyperlink = f"https://www.liepin.com/company/{cid}/"
            if col == 6:
                jid = j.get("encrypt_job_id", "")
                if jid: c.font = link_font; c.hyperlink = f"https://www.liepin.com/job/{jid}.shtml"
        row += 1

    for i, w in enumerate([6, 12, 16, 22, 24, 28, 14, 16], 1):
        ws.column_dimensions[chr(64 + i)].width = w

    basename = f"猎聘_投递_{today}"
    filepath = os.path.join(liepin_dir, f"{basename}.xlsx")
    counter = 1
    while os.path.exists(filepath):
        filepath = os.path.join(liepin_dir, f"{basename}({counter}).xlsx")
        counter += 1
    wb.save(filepath)
    return f"✅ 猎聘投递 Excel 已生成：`{filepath}`（共 {len(jobs)} 条）"


def export_liepin_recommend_excel(date_filter: str = None) -> str:
    """只导出猎聘推荐"""
    from storage import _load_jobs
    from datetime import date as _date
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    liepin_dir = os.path.join(DATA_DIR, "reports", "liepin")
    os.makedirs(liepin_dir, exist_ok=True)

    if date_filter:
        today = date_filter
    else:
        today = _date.today().isoformat()

    jobs = _load_jobs()
    jobs = [j for j in jobs if j.get("platform") == "猎聘" and j.get("status") == "猎聘-推荐"]
    if not jobs:
        return f"暂无猎聘推荐数据。"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "猎聘推荐"
    hfont = Font(bold=True, size=11)
    hfill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    bdr = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    link_font = Font(color="0563C1", underline="single")

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    ws.cell(row=row, column=1, value="猎聘 — 职位推荐").font = Font(bold=True, size=13, color="1F4E79")
    row += 1

    for col, h in enumerate(["序号", "类型", "企业", "招聘职位", "薪资", "城市", "经验", "学历", "规模"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = hfont; c.fill = hfill; c.border = bdr; c.alignment = Alignment(horizontal="center")
    row += 1

    for i, j in enumerate(jobs):
        vals = [i + 1, "📋 推荐", j.get("company", ""), j.get("position", ""), j.get("salary", ""),
                j.get("city", ""), j.get("experience", ""), j.get("degree", ""), j.get("scale", "")]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = bdr; c.alignment = Alignment(vertical="center")
            if col == 3:
                cid = j.get("encrypt_brand_id", "")
                if cid: c.font = link_font; c.hyperlink = f"https://www.liepin.com/company/{cid}/"
            if col == 4:
                jid = j.get("encrypt_job_id", "")
                if jid: c.font = link_font; c.hyperlink = f"https://www.liepin.com/job/{jid}.shtml"
        row += 1

    for i, w in enumerate([6, 10, 22, 30, 14, 10, 10, 8, 20], 1):
        ws.column_dimensions[chr(64 + i)].width = w

    basename = f"猎聘_推荐_{today}"
    filepath = os.path.join(liepin_dir, f"{basename}.xlsx")
    counter = 1
    while os.path.exists(filepath):
        filepath = os.path.join(liepin_dir, f"{basename}({counter}).xlsx")
        counter += 1
    wb.save(filepath)
    return f"✅ 猎聘推荐 Excel 已生成：`{filepath}`（共 {len(jobs)} 条）"
